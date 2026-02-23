"""Excel export using openpyxl."""

import io
import logging

logger = logging.getLogger(__name__)


def export_to_excel(report_data: dict) -> bytes:
    """Convert report data dict to an Excel file (.xlsx) in memory."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = report_data.get("title", "Report")[:31]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(report_data.get("columns", [])), 3))
    title_cell = ws.cell(row=1, column=1, value=report_data.get("title", "Report"))
    title_cell.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Period: {report_data.get('period', '')}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=10)

    # Summary section
    summary = report_data.get("summary", {})
    row_num = 4
    for key, value in summary.items():
        if isinstance(value, dict):
            ws.cell(row=row_num, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=str(value))
        else:
            ws.cell(row=row_num, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=value)
        row_num += 1

    # Data table
    row_num += 1
    columns = report_data.get("columns", [])
    rows = report_data.get("rows", [])

    # Column headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row_num += 1

    # Data rows
    for row_data in rows:
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            if isinstance(value, dict):
                value = str(value)
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1

    # Auto-adjust column widths
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
